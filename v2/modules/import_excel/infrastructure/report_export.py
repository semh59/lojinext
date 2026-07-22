"""
TYPE: SINGLETON
SCOPE: Application lifetime
SINGLETON_REASON: Dışa aktarma servisi — Excel/PDF üretimi, stateless.
CREATED_BY: v2/modules/platform_infra/container.py (lazy property)
"""

import os
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Dict, Optional

# openpyxl
try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from app.infrastructure.logging.logger import get_logger
from v2.modules.reports.public import get_report_generator

logger = get_logger(__name__)

_FORMULA_STARTERS = frozenset("=+-@|%")


def _safe_cell(v: object) -> str:
    """Convert value to string and strip leading formula starters (CSV/Excel injection)."""
    s = str(v) if v is not None else "-"
    return "'" + s if s and s[0] in _FORMULA_STARTERS else s


class ExportService:
    """
    Export Servisi (Async & Unified PDF)

    Desteklenen formatlar:
    - Excel (.xlsx) - openpyxl
    - PDF (.pdf) - ReportLab
    """

    # Cross-platform export directory
    # Windows: %APPDATA%/LojiNext/exports
    # Linux/Mac: ~/.lojinext/exports
    @staticmethod
    def _get_export_dir() -> Path:
        if os.name == "nt":  # Windows
            base = Path(os.getenv("APPDATA", Path.home()))
        else:  # Linux/Mac
            base = Path.home() / ".lojinext"
        return base / "LojiNext" / "exports"

    EXPORT_DIR = _get_export_dir()

    def _sanitize_filename(self, filename: str) -> str:
        """Zararlı karakterleri temizler ve path traversal'ı engeller (Path Traversal Guard)"""
        import re

        # Sadece güvenli karakterlere izin ver (Alfanümerik, nokta, tire, alt çizgi)
        safe_filename = re.sub(r"[^a-zA-Z0-9._-]", "_", filename)
        # Path traversal engellemek için sadece dosya ismini al
        return os.path.basename(safe_filename)

    def cleanup_old_exports(self, max_age_days: int = 7) -> int:
        """EXPORT_DIR içindeki max_age_days gün öncesi dosyaları siler. Silinen sayı döner."""
        from datetime import timedelta

        cutoff = datetime.now(timezone.utc) - timedelta(days=max_age_days)
        removed = 0
        try:
            if not self.EXPORT_DIR.exists():
                return 0
            for path in self.EXPORT_DIR.iterdir():
                if not path.is_file():
                    continue
                try:
                    mtime = datetime.fromtimestamp(
                        path.stat().st_mtime, tz=timezone.utc
                    )
                    if mtime < cutoff:
                        path.unlink()
                        removed += 1
                except OSError:
                    pass
        except OSError as exc:
            logger.warning("cleanup_old_exports failed: %s", exc)
        return removed

    # =========================================================================
    # EXCEL EXPORT
    # =========================================================================

    async def export_to_excel(
        self, data: Dict, filename: str, title: str = "Rapor"
    ) -> Optional[str]:
        """Excel export (Async & Non-blocking)"""
        import asyncio

        self.cleanup_old_exports()
        filename = self._sanitize_filename(filename)
        return await asyncio.to_thread(
            self._export_to_excel_sync, data, filename, title
        )

    def _export_to_excel_sync(
        self, data: Dict, filename: str, title: str
    ) -> Optional[str]:
        """Senkron Excel üretimi (Internal)"""
        if not OPENPYXL_AVAILABLE:
            logger.error("openpyxl yüklü değil!")
            return None

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = title[:31]

            # Başlık ve Tarih
            ws.merge_cells("A1:F1")
            ws["A1"] = _safe_cell(title)
            ws["A1"].font = Font(bold=True, size=16)
            ws["A1"].alignment = Alignment(horizontal="center")
            ws["A2"] = (
                f"Oluşturma: {datetime.now(timezone.utc).strftime('%d.%m.%Y %H:%M')}"
            )

            row = 4
            for section, content in data.items():
                if not content:
                    continue

                ws.cell(
                    row=row, column=1, value=_safe_cell(section.upper())
                ).font = Font(bold=True)
                row += 1

                if (
                    isinstance(content, list)
                    and content
                    and isinstance(content[0], dict)
                ):
                    headers = list(content[0].keys())
                    for col, h in enumerate(headers, 1):
                        cell = ws.cell(row=row, column=col, value=_safe_cell(h))
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill("solid", fgColor="1E40AF")
                    row += 1
                    for item in content:
                        for col, k in enumerate(headers, 1):
                            ws.cell(
                                row=row, column=col, value=_safe_cell(item.get(k, "-"))
                            )
                        row += 1
                elif isinstance(content, dict):
                    for k, v in content.items():
                        ws.cell(row=row, column=1, value=_safe_cell(k))
                        ws.cell(row=row, column=2, value=_safe_cell(v))
                        row += 1
                row += 1

            if not filename.endswith(".xlsx"):
                filename += ".xlsx"
            filepath = self.EXPORT_DIR / filename
            filepath.parent.mkdir(parents=True, exist_ok=True)
            wb.save(filepath)
            return str(filepath)
        except Exception as e:
            logger.error(f"Excel export error: {e}")
            return None

    # =========================================================================
    # PDF EXPORT (Async & Non-blocking)
    # =========================================================================

    async def export_fleet_summary_pdf(
        self, start_date: date, end_date: date, data: Dict, filename: str
    ) -> Optional[str]:
        """Filo özetini PDF olarak kaydet (Async & Non-blocking)"""
        import asyncio

        try:
            filename = self._sanitize_filename(filename)
            generator = get_report_generator()
            # Asenkron wrapper'ı kullan (zaten to_thread içerir)
            pdf_bytes = await generator.async_generate_fleet_summary(
                start_date, end_date, data
            )

            if not filename.endswith(".pdf"):
                filename += ".pdf"
            filepath = self.EXPORT_DIR / filename

            # I/O yazma işlemini thread'e al
            def _write_file(path, data):
                with open(path, "wb") as f:
                    f.write(data)

            await asyncio.to_thread(_write_file, filepath, pdf_bytes)
            return str(filepath)
        except Exception as e:
            logger.error(f"PDF Fleet Export Error: {e}")
            return None

    async def export_vehicle_report_pdf(
        self, arac_id: int, month: int, year: int, data: Dict, filename: str
    ) -> Optional[str]:
        """Araç raporunu PDF olarak kaydet (Async & Non-blocking)"""
        import asyncio

        try:
            filename = self._sanitize_filename(filename)
            generator = get_report_generator()
            # Asenkron wrapper'ı kullan
            pdf_bytes = await generator.async_generate_vehicle_report(
                arac_id, month, year, data
            )

            if not filename.endswith(".pdf"):
                filename += ".pdf"
            filepath = self.EXPORT_DIR / filename

            def _write_file(path, data):
                with open(path, "wb") as f:
                    f.write(data)

            await asyncio.to_thread(_write_file, filepath, pdf_bytes)
            return str(filepath)
        except Exception as e:
            logger.error(f"PDF Vehicle Export Error: {e}")
            return None

    # =========================================================================
    # EXCEL TEMPLATE GENERATION (NEW Bölüm 6)
    # =========================================================================

    async def generate_template(self, entity_type: str) -> Optional[str]:
        """Modellerle tam uyumlu Excel şablonu üret (Async)"""
        import asyncio

        return await asyncio.to_thread(self._generate_template_sync, entity_type)

    def _generate_template_sync(self, entity_type: str) -> Optional[str]:
        """Senkron Şablon Üretimi (Internal)"""
        if not OPENPYXL_AVAILABLE:
            return None

        templates: Dict[str, Dict[str, list]] = {
            "yakit": {
                "columns": [
                    "tarih",
                    "plaka",
                    "litre",
                    "tutar",
                    "km",
                    "istasyon",
                    "fis_no",
                    "depo_durumu",
                ],
                "sample": [
                    "2024-01-01",
                    "34ABC123",
                    450.5,
                    15000.0,
                    125400,
                    "Opet Gebze",
                    "A123",
                    "Dolu",
                ],
            },
            "sefer": {
                "columns": [
                    "tarih",
                    "sofor",
                    "plaka",
                    "cikis",
                    "varis",
                    "km",
                    "ton",
                    "saat",
                ],
                "sample": [
                    "2024-01-01",
                    "Ahmet Yılmaz",
                    "34ABC123",
                    "İstanbul",
                    "Ankara",
                    450,
                    22.5,
                    "08:00",
                ],
            },
            "arac": {
                "columns": [
                    "plaka",
                    "marka",
                    "model",
                    "yil",
                    "tank_kapasitesi",
                    "bos_agirlik_kg",
                    "motor_verimliligi",
                ],
                "sample": ["34ABC123", "Mercedes", "Actros", 2022, 600, 8200, 0.38],
            },
            "sofor": {
                "columns": ["ad_soyad", "telefon", "ise_baslama", "ehliyet_sinifi"],
                "sample": ["Ahmet Yılmaz", "05321234455", "2023-05-15", "E"],
            },
        }

        if entity_type not in templates:
            logger.error(f"Geçersiz şablon tipi: {entity_type}")
            return None

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = f"{entity_type.capitalize()} Sablonu"

            t_data = templates[entity_type]
            cols = t_data["columns"]

            header_fill = PatternFill(
                start_color="1E40AF", end_color="1E40AF", fill_type="solid"
            )
            header_font = Font(color="FFFFFF", bold=True)
            border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            for col_idx, col_name in enumerate(cols, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                ws.column_dimensions[get_column_letter(col_idx)].width = 20

                sample_cell = ws.cell(
                    row=2, column=col_idx, value=t_data["sample"][col_idx - 1]
                )
                sample_cell.border = border

            filename = self._sanitize_filename(f"{entity_type}_sablon.xlsx")
            filepath = self.EXPORT_DIR / filename
            filepath.parent.mkdir(parents=True, exist_ok=True)
            wb.save(filepath)

            return str(filepath)
        except Exception as e:
            logger.error(f"Şablon üretim hatası: {e}")
            return None


def get_export_service() -> ExportService:
    """Delegates to the DI container for the singleton ExportService instance."""
    from v2.modules.platform_infra.container import get_container

    return get_container().export_service
